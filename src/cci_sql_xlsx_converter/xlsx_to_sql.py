import argparse
import openpyxl # Used for reading from the XLSX file
import json # Used to parse metadata in cell comments and file description, because it doesn't rely on whitespace like YAML
import yaml # Used for saving CCI mapping data saved in the XLSX file's description
import os, platform # Used for opening the generated XLSX file
import send2trash # Used for deleting the XLSX file, so the user can get it back if this was done by accident

def xlsx_to_sql():
  # Parse command line arguments
  parser = argparse.ArgumentParser()
  parser.add_argument("-i", "--input", help="Path to the XLSX script to convert", default="./generated.xlsx")
  parser.add_argument("-o", "--output-sql", help="Path to the SQL file to create", default="./datasets/sample.sql")
  parser.add_argument("-y", "--output-yaml", help="Path to the YAML mapping file to create", default="./datasets/mapping.yml")

  parser.add_argument("-ltn", "--log-table-names", help="Log names of detected tables", action="store_true")
  parser.add_argument("-lr", "--log-records", help="Log detected records", action="store_true")
  parser.add_argument("-lf", "--log-fields", help="Log names of detected fields in tables", action="store_true")

  parser.add_argument("-sw", "--suppress-warnings", help="Do not display warnings", action="store_true")

  output_options = parser.add_argument_group("XLSX file options")
  parser.add_argument("-d", "--delete-xlsx", help="Do not delete the XLSX file", action="store_true")
  parser.add_argument("-p", "--preserve-xlsx", help="Do not delete the XLSX file", action="store_false", dest='delete_xlsx')
  # By default, delete the file at the end, so that there aren't 2 versions of the data
  parser.set_defaults(delete_xlsx=True)

  output_options = parser.add_argument_group("Output options")
  output_options.add_argument("-of", "--open-file", help="Automatically open the generated SQL file", action="store_true")
  output_options.add_argument("-nof", "--no-open-file", help="Do not open the generated SQL file", action="store_false", dest='open_file')
  # By default, don't open the file at the end: assume the user will manually do that when using the file in CCI
  parser.set_defaults(open_file=False)

  args = parser.parse_args()

  # Output options
  LOG_TABLE_NAMES = args.log_table_names
  LOG_RECORDS = args.log_records
  LOG_FIELDS = args.log_fields

  DISPLAY_WARNINGS = not args.suppress_warnings

  OPEN_FILE = args.open_file
  DELETE_XLSX = args.delete_xlsx

  INPUT_FILE = args.input
  OUTPUT_SQL_FILE = args.output_sql
  OUTPUT_YAML_FILE = args.output_yaml

  COMMENT_PREFIX = "AUTOGENERATED, DO NOT EDIT!\n" # Must be kept identical with COMMENT_PREFIX in the other script
  DEFAULT_METADATA = {"type": "VARCHAR(255)", "not_null": False, "pk": 0} # Used for fields with no comment, assumed to be added by the user in Excel

  # Logging functions
  def log_table_name(text):
    if LOG_TABLE_NAMES: print(text)
  def log_record(text):
    if LOG_RECORDS: print(text)
  def log_field(text):
    if LOG_FIELDS: print(text)
  def log_warning(text):
    if DISPLAY_WARNINGS: print(text)

  class XLSXParseError(Exception): pass

  # Load the XLSX file
  workbook = openpyxl.load_workbook(INPUT_FILE)

  # Parse YAML mapping data
  mapping_original = json.loads(workbook.properties.description)
  # Extract list of tables without mapping data, then remove it from the mapping object
  tables_without_mapping_data = mapping_original.pop("__cci_sql_xlsx_converter_tables_without_mapping_data__", None)
  if tables_without_mapping_data is None:
    raise XLSXParseError("Mapping data does not contain list of tables without mapping data (this list must be present even if empty)")

  # Create a dictionary recording whether each table's mapping data has been used.
  # This is so we can print a warning if a table's mapping data is not used.
  # This could just mean the table was deleted in the XLSX file, but it also could mean
  # the user manually edited the table's name. Either way, we can't recover the mapping data,
  # so we print a warning.
  table_mappings_preserved = {block_name : False for block_name in mapping_original.keys()}
  mapping_new = {}

  # Variable to hold contents of the SQL script
  sql_script = "BEGIN TRANSACTION;\n"

  # For each worksheet, create the corresponding table in the SQL script
  for sheet in workbook.worksheets:
    table_name = sheet.title
    log_table_name(f"Extracting table: {table_name}")

    # Collect list of fields in this table
    fields = []
    mapping_fields = [] # Fields which are not the primary key
    field_metadata = []
    column = 1
    cell = sheet.cell(row=1, column=column)
    while cell.value:
      field_name = cell.value
      fields.append(field_name)

      # Extract field metadata
      if cell.comment:
        comment_text = cell.comment.text
        prefix = comment_text[:len(COMMENT_PREFIX)]
        suffix = comment_text[len(COMMENT_PREFIX):]

        if prefix != COMMENT_PREFIX:
          raise XLSXParseError(f"Comment verification failed: '{comment_text}'")

        metadata = json.loads(suffix)
        field_metadata.append(metadata)
      else:
        log_warning(f"WARNING: No comment found for field {field_name}. Using default type: VARCHAR(255), not primary key.")
        metadata = DEFAULT_METADATA
        field_metadata.append(metadata)

      column += 1
      cell = sheet.cell(row=1, column=column)
      log_field(f"Detected field: {field_name} with metadata {metadata}")
    log_field(f"Detected {len(fields)} fields. The final, empty cell's value's type was: {str(type(cell.value))}")

    # Create the table
    sql_script += f'CREATE TABLE "{table_name}" (\n'

    primary_key_field = None

    # Add the fields
    for column, field_name in enumerate(fields):
      metadata = field_metadata[column]
      if metadata["pk"]:
        if primary_key_field is not None:
          raise XLSXParseError(f"Table {table_name}: multiple primary key fields detected!")
        
        primary_key_field = field_name

      # Primary key field doesn't get quotes around its name
      if primary_key_field == field_name:
        sql_script += f'\t{field_name}'
      else:
        sql_script += f'\t"{field_name}"'
        mapping_fields.append(field_name)

      sql_script += f' {metadata["type"]}'
      if metadata["not_null"]: sql_script += " NOT NULL"
      sql_script += ",\n"

    # Add the primary key
    if primary_key_field is not None:
      sql_script += f'  PRIMARY KEY ({primary_key_field})\n'
    else:
      raise XLSXParseError(f"Table {table_name}: no primary key field detected!")

    # End the table with close-paren
    sql_script += ");\n"

    # Insert records into the table
    for row_index, row_obj in enumerate(sheet.rows):
      if row_index == 0: continue # Skip the header row

      # Skip empty rows
      if not any(cell.value for cell in row_obj):
        continue

      values = []
      for column_index, field in enumerate(fields):
        value = sheet.cell(row=row_index+1, column=column_index+1).value # Remember, 1-based indexing!
        # CCI uses an empty string to indicate NULL, but OpenPyXL uses None
        if value is None: value = ''

        target_type = field_metadata[column_index]["type"]
        if target_type == "INTEGER":
          value = int(value)
        else:
          value = str(value) # Ensure the value is a string, since Excel likes to auto-format booleans and numbers

        values.append(value)

      sql_script += f'INSERT INTO "{table_name}" VALUES('
      for column, value in enumerate(values):
        if column > 0: sql_script += ","
        if isinstance(value, str): sql_script += f"'{value}'"
        else: sql_script += f"{value}"
      sql_script += ");\n"

      log_record(f"Record in table {table_name}: {values}")

    # Build new mapping data
    block_name = 'Insert ' + table_name
    if block_name in mapping_original.keys():
      table_mappings_preserved[block_name] = True
      mapping_new[block_name] = mapping_original[block_name]
      mapping_new[block_name]['fields'] = mapping_fields
    elif table_name in tables_without_mapping_data:
      pass # Allow the new mapping data to lack a block for this table, since there was none in the original mapping
    else:
      log_warning(f"WARNING: Mapping data for table {table_name} not found. Adding default data.")
      mapping_new[block_name] = {'table': table_name, 'fields': mapping_fields}

  # End the SQL file
  sql_script += "COMMIT;\n"

  # Warn about unused mapping data
  for block_name in mapping_original.keys():
    if not table_mappings_preserved[block_name]:
      log_warning(f"WARNING: Mapping data for block {block_name} not found. It will not be included in the SQL file.\nIf you deleted the corresponding table, ignore this. If you changed the table's name, you'll have to restore the following mapping data (excluding fields):\n{mapping_original[block_name]}\n")

  # Save the SQL file
  with open(OUTPUT_SQL_FILE, 'w') as f:
    f.write(sql_script)

  # Save the YAML file
  with open(OUTPUT_YAML_FILE, 'w') as f:
    yaml.dump(mapping_new, f)

  # Delete the XLSX file
  if DELETE_XLSX:
    send2trash.send2trash(INPUT_FILE)

  # Open the created SQL file
  if OPEN_FILE:
    if platform.system() == "Windows":
      os.startfile(OUTPUT_SQL_FILE)
    else:
      os.system(f"open {OUTPUT_SQL_FILE}")

if __name__ == "__main__":
  xlsx_to_sql()